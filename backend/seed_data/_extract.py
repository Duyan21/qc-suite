"""One-off extraction script: docs/Home_lending/*.xlsx -> backend/seed_data/*.json.

Not imported by the app or by seed.py. Re-run manually if the source
spreadsheets in docs/Home_lending change; requires openpyxl.

    pip install openpyxl
    python backend/seed_data/_extract.py
"""
import json
import os

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DOCS_DIR = os.path.join(ROOT, "docs", "Home_lending")
OUT_DIR = os.path.dirname(os.path.abspath(__file__))

MODULES = [
    "Customer & Loan Application",
    "Customer Eligibility & Verification",
    "Property Assessment",
    "Credit Assessment & Approval",
    "Contract & Disbursement",
    "Loan Servicing",
]

# Hand-authored descriptions grounded in the BRS process/business rules
# (BRS has titles only; these give each requirement real, one-sentence context).
REQ_DESCRIPTIONS = {
    "REQ-001": "Customer creates a new home loan application by providing personal details, employment information, and the desired loan amount and property.",
    "REQ-002": "Customer saves an in-progress application as a draft to complete later without losing entered data.",
    "REQ-003": "Customer resumes a previously saved draft application and continues from where they left off.",
    "REQ-004": "Customer updates details of an application that has not yet been submitted for review.",
    "REQ-005": "Customer uploads supporting documents (ID, payslips, bank statements, property papers) required for the application.",
    "REQ-006": "Customer deletes a draft application that is no longer needed.",
    "REQ-007": "Customer submits a completed application for processing; KYC must be completed before submission is allowed.",
    "REQ-008": "Customer views the current status of a submitted application as it moves through the approval pipeline.",
    "REQ-009": "System verifies the customer's identity against submitted ID documents before proceeding with eligibility checks.",
    "REQ-010": "System validates that the applicant meets the minimum applicant age of 18 years.",
    "REQ-011": "System validates the applicant's citizenship/residency status against program eligibility rules.",
    "REQ-012": "System verifies the applicant's employment status and employer details.",
    "REQ-013": "System verifies the applicant's income using payslips or bank statements issued within the last 6 months.",
    "REQ-014": "System reviews the applicant's existing debt obligations as an input to the credit assessment.",
    "REQ-015": "System screens the applicant against anti-money-laundering (AML) sanctioned-party lists before KYC approval.",
    "REQ-016": "System finalizes KYC approval once identity, age, citizenship, and AML checks have all passed.",
    "REQ-017": "Customer adds the property being purchased or used as collateral to the application.",
    "REQ-018": "System verifies legal ownership of the property against the submitted title/ownership documents.",
    "REQ-019": "System validates the property address for completeness and correctness.",
    "REQ-020": "System validates that the property type is eligible as collateral under lending policy (e.g., agricultural land is excluded).",
    "REQ-021": "System records the appraised property valuation, valid for 90 days from the assessment date.",
    "REQ-022": "System calculates the loan-to-value (LTV) ratio and enforces the maximum LTV of 80%.",
    "REQ-023": "System validates that required property insurance is in place before property approval.",
    "REQ-024": "System finalizes property approval once ownership, valuation, and insurance checks pass.",
    "REQ-025": "System assesses the applicant's credit score to support the lending decision.",
    "REQ-026": "System calculates the debt-to-income (DTI) ratio and enforces the maximum DTI of 40%.",
    "REQ-027": "System calculates the maximum loan amount the applicant qualifies for based on income, DTI, and LTV limits.",
    "REQ-028": "System calculates the applicable interest rate, including any eligible promotional rates.",
    "REQ-029": "System generates the repayment schedule based on loan amount, interest rate, and tenure (maximum 35 years).",
    "REQ-030": "System routes borderline applications to a loan officer for manual review.",
    "REQ-031": "System supports conditional approval of an application pending outstanding conditions.",
    "REQ-032": "Branch manager approves the loan application once all assessment criteria are satisfied.",
    "REQ-033": "System records the rejection of an application that fails eligibility, credit, or policy checks.",
    "REQ-034": "System generates the loan contract document based on approved loan terms.",
    "REQ-035": "Customer digitally signs the loan contract using a valid, non-expired certificate.",
    "REQ-036": "Customer accepts the loan terms and conditions prior to disbursement.",
    "REQ-037": "System schedules the disbursement of loan funds after contract signing.",
    "REQ-038": "System executes a single, idempotent transfer of loan funds to the customer.",
    "REQ-039": "System notifies the customer of key contract and disbursement milestones.",
    "REQ-040": "System activates the loan only after disbursement has been confirmed.",
    "REQ-041": "System confirms successful disbursement and records the confirmation.",
    "REQ-042": "Customer views the current repayment schedule for an active loan.",
    "REQ-043": "System processes a scheduled repayment against an active loan.",
    "REQ-044": "System accepts a partial repayment and recalculates the outstanding balance.",
    "REQ-045": "System calculates the payoff amount, including any early settlement penalty, using the currently effective rate.",
    "REQ-046": "System calculates penalties for late or early repayment events.",
    "REQ-047": "System generates a loan statement showing repayment history and outstanding balance.",
    "REQ-048": "Customer updates personal/contact details on their profile.",
    "REQ-049": "System closes a loan once fully repaid; a closed loan must no longer accept repayments.",
    "REQ-050": "Customer views the history of notifications (SMS/email) sent regarding their loan, without duplicates.",
}

REQ_TITLES = {
    "REQ-001": "Create Home Loan Application", "REQ-002": "Save Draft",
    "REQ-003": "Resume Draft", "REQ-004": "Update Application",
    "REQ-005": "Upload Supporting Documents", "REQ-006": "Delete Draft",
    "REQ-007": "Submit Application", "REQ-008": "Track Application Status",
    "REQ-009": "Customer Identity Verification", "REQ-010": "Age Validation",
    "REQ-011": "Citizenship Validation", "REQ-012": "Employment Verification",
    "REQ-013": "Income Verification", "REQ-014": "Existing Debt Assessment",
    "REQ-015": "AML Screening", "REQ-016": "KYC Approval",
    "REQ-017": "Add Property", "REQ-018": "Property Ownership Verification",
    "REQ-019": "Property Address Validation", "REQ-020": "Property Type Validation",
    "REQ-021": "Property Valuation", "REQ-022": "Loan-to-Value Calculation",
    "REQ-023": "Insurance Validation", "REQ-024": "Property Approval",
    "REQ-025": "Credit Score Assessment", "REQ-026": "Debt-to-Income Calculation",
    "REQ-027": "Maximum Loan Calculation", "REQ-028": "Interest Rate Calculation",
    "REQ-029": "Repayment Schedule Calculation", "REQ-030": "Manual Review",
    "REQ-031": "Conditional Approval", "REQ-032": "Loan Approval",
    "REQ-033": "Loan Rejection", "REQ-034": "Generate Loan Contract",
    "REQ-035": "Digital Signature", "REQ-036": "Accept Terms",
    "REQ-037": "Schedule Disbursement", "REQ-038": "Execute Disbursement",
    "REQ-039": "Customer Notification", "REQ-040": "Loan Activation",
    "REQ-041": "Disbursement Confirmation", "REQ-042": "View Repayment Schedule",
    "REQ-043": "Repayment Processing", "REQ-044": "Partial Repayment",
    "REQ-045": "Early Settlement", "REQ-046": "Penalty Calculation",
    "REQ-047": "Loan Statement", "REQ-048": "Update Customer Profile",
    "REQ-049": "Close Loan", "REQ-050": "Notification History",
}

REQ_MODULE_RANGES = [
    (1, 8, MODULES[0]), (9, 16, MODULES[1]), (17, 24, MODULES[2]),
    (25, 33, MODULES[3]), (34, 41, MODULES[4]), (42, 50, MODULES[5]),
]


def module_for(n):
    for lo, hi, m in REQ_MODULE_RANGES:
        if lo <= n <= hi:
            return m
    raise ValueError(n)


def build_requirements():
    reqs = []
    for n in range(1, 51):
        req_id = f"REQ-{n:03d}"
        reqs.append({
            "req_id": req_id,
            "module": module_for(n),
            "title": REQ_TITLES[req_id],
            "description": REQ_DESCRIPTIONS[req_id],
        })
    return reqs


def build_test_cases():
    cases = []
    for m in range(1, 7):
        path = os.path.join(DOCS_DIR, f"Home_Lending_Test_Cases_Module{m}.xlsx")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header, data_rows = rows[0], rows[1:]
        for row in data_rows:
            (tc_id, req_id, module, scenario, preconditions, steps,
             test_data, expected_result, priority, test_type, status) = row
            cases.append({
                "tc_id": tc_id,
                "req_id": req_id,
                "module": module,
                "scenario": scenario,
                "preconditions": preconditions,
                "steps": steps,
                "test_data": test_data,
                "expected_result": expected_result,
                "priority": priority,
                "test_type": test_type,
                "status": status,
            })
    return cases


def build_defects():
    path = os.path.join(DOCS_DIR, "Home_Lending_Defect_Log_Polished.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    defects = []
    for row in data_rows:
        (def_id, summary, module, req_id, tc_id, severity, priority, environment,
         steps_to_reproduce, expected_result, actual_result, root_cause,
         category, status) = row
        defects.append({
            "def_id": def_id,
            "summary": summary,
            "module": module,
            "req_id": req_id,
            "tc_id": tc_id,
            "severity": severity,
            "priority": priority,
            "environment": environment,
            "steps_to_reproduce": steps_to_reproduce,
            "expected_result": expected_result,
            "actual_result": actual_result,
            "root_cause": root_cause,
            "category": category,
            "status": status,
        })
    return defects


def main():
    requirements = build_requirements()
    test_cases = build_test_cases()
    defects = build_defects()

    with open(os.path.join(OUT_DIR, "requirements.json"), "w", encoding="utf-8") as f:
        json.dump(requirements, f, indent=2, ensure_ascii=False)
    with open(os.path.join(OUT_DIR, "test_cases.json"), "w", encoding="utf-8") as f:
        json.dump(test_cases, f, indent=2, ensure_ascii=False)
    with open(os.path.join(OUT_DIR, "defects.json"), "w", encoding="utf-8") as f:
        json.dump(defects, f, indent=2, ensure_ascii=False)

    print(f"requirements: {len(requirements)}, test_cases: {len(test_cases)}, defects: {len(defects)}")


if __name__ == "__main__":
    main()
